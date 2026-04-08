import base64
import io
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning('openpyxl no instalado. Instalar con: pip install openpyxl')
    openpyxl = None


class ImportProductosWizard(models.TransientModel):
    _name = 'ferreteria.import.productos.wizard'
    _description = 'Importar Productos desde Excel'

    file = fields.Binary(
        string='Archivo Excel',
        required=True,
        help='Archivo Excel (.xlsx) con lista de precios',
    )
    filename = fields.Char(string='Nombre del Archivo')
    update_existing = fields.Boolean(
        string='Actualizar productos existentes',
        default=True,
        help='Si está activo, actualiza precios de productos que ya existen (por código)',
    )
    import_mode = fields.Selection([
        ('all', 'Todos los productos'),
        ('with_price', 'Solo con precio definido'),
    ], string='Modo de Importación', default='all')

    # Resultado
    result_message = fields.Text(string='Resultado', readonly=True)
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('done', 'Completado'),
    ], default='draft')

    def _parse_price(self, value):
        """Convierte un valor de precio del Excel a float.
        Maneja formatos: 'S/ 17,50', '17.50', 17.5, '-', None
        """
        if not value or str(value).strip() in ('-', '', '0', '0.0', '0,0'):
            return 0.0
        text = str(value).strip()
        # Quitar prefijo de moneda
        text = text.replace('S/', '').replace('s/', '').strip()
        # Formato peruano: punto = miles, coma = decimal
        # Si tiene coma, asumir formato "1.234,56" o "17,50"
        if ',' in text:
            text = text.replace('.', '').replace(',', '.')
        try:
            return float(text)
        except ValueError:
            return 0.0

    def _get_departamento_categoria(self, departamento):
        """Mapea departamento del Excel a categoría de ferretería."""
        mapping = {
            'FERRETERIA': 'ferreteria_inventario.cat_otros',
            'GASFITERIA': 'ferreteria_inventario.cat_gasfiteria',
            'HERRAMIENTAS': 'ferreteria_inventario.cat_herramientas',
            'PINTURA Y ACABADOS': 'ferreteria_inventario.cat_pinturas',
            'ELECTRICIDAD': 'ferreteria_inventario.cat_electricidad',
            'EPPS': 'ferreteria_inventario.cat_seguridad',
            'LIMPIEZA': 'ferreteria_inventario.cat_otros',
            'SIN DEFINIR': False,
        }
        xml_id = mapping.get(str(departamento).strip().upper(), False)
        if xml_id:
            return self.env.ref(xml_id, raise_if_not_found=False)
        return False

    def action_import(self):
        """Ejecuta la importación de productos desde el archivo Excel."""
        self.ensure_one()

        if not openpyxl:
            raise UserError(_(
                'La librería openpyxl no está instalada. '
                'Ejecute: pip install openpyxl'
            ))

        if not self.file:
            raise UserError(_('Debe seleccionar un archivo Excel.'))

        # Leer archivo
        file_data = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(file_data),
            data_only=True,
        )

        ws = wb.active
        if not ws:
            raise UserError(_('El archivo Excel no tiene hojas.'))

        ProductTemplate = self.env['product.template']
        Pricelist = self.env['product.pricelist']
        PricelistItem = self.env['product.pricelist.item']

        # Obtener las 4 listas de precios
        pricelist_refs = [
            'ferreteria_inventario.pricelist_mayoreo1',
            'ferreteria_inventario.pricelist_mayoreo2',
            'ferreteria_inventario.pricelist_mayoreo3',
            'ferreteria_inventario.pricelist_mayoreo4',
        ]
        pricelists = []
        for ref in pricelist_refs:
            pl = self.env.ref(ref, raise_if_not_found=False)
            if pl:
                pricelists.append(pl)

        created = 0
        updated = 0
        skipped = 0
        errors = []
        current_departamento = 'SIN DEFINIR'
        current_proveedor = 'SIN DEFINIR'

        # Recorrer filas (empezar desde fila 6 que es donde comienzan datos)
        for row_idx in range(6, ws.max_row + 1):
            try:
                a_val = ws.cell(row=row_idx, column=1).value
                d_val = ws.cell(row=row_idx, column=4).value
                f_val = ws.cell(row=row_idx, column=6).value
                i_val = ws.cell(row=row_idx, column=9).value
                l_val = ws.cell(row=row_idx, column=12).value
                m_val = ws.cell(row=row_idx, column=13).value
                p_val = ws.cell(row=row_idx, column=16).value

                # Fila de metadatos (departamento/proveedor)
                if f_val == 'Cantidad':
                    if a_val:
                        current_departamento = str(a_val).strip()
                    if d_val:
                        current_proveedor = str(d_val).strip()
                    continue

                # Fila sin código = saltar
                if not a_val:
                    continue

                code = str(a_val).strip()

                # Saltar filas de encabezado
                if code in ('Clave', ' Clave', 'Departamento:'):
                    continue

                description = str(d_val).strip() if d_val else ''
                if not description:
                    continue

                # Parsear precios
                prices = [
                    self._parse_price(i_val),
                    self._parse_price(l_val),
                    self._parse_price(m_val),
                    self._parse_price(p_val),
                ]

                # Modo: solo con precio
                if self.import_mode == 'with_price' and prices[0] == 0.0:
                    skipped += 1
                    continue

                # Buscar categoría
                categoria = self._get_departamento_categoria(current_departamento)

                # Buscar producto existente por código
                existing = ProductTemplate.search([
                    ('default_code', '=', code),
                ], limit=1)

                if existing:
                    if self.update_existing:
                        vals = {
                            'list_price': prices[0] if prices[0] > 0 else existing.list_price,
                        }
                        if categoria:
                            vals['ferreteria_categoria_id'] = categoria.id
                        existing.write(vals)
                        product = existing
                        updated += 1
                    else:
                        skipped += 1
                        continue
                else:
                    # Crear producto nuevo
                    vals = {
                        'name': description,
                        'default_code': code,
                        'list_price': prices[0] if prices[0] > 0 else 0.0,
                        'type': 'product',
                        'es_ferreteria': True,
                        'sale_ok': True,
                        'purchase_ok': True,
                    }
                    if categoria:
                        vals['ferreteria_categoria_id'] = categoria.id
                    if current_proveedor and current_proveedor != 'SIN DEFINIR':
                        vals['marca'] = current_proveedor

                    product = ProductTemplate.create(vals)
                    created += 1

                # Configurar precios en las listas
                for idx, pricelist in enumerate(pricelists):
                    price = prices[idx] if idx < len(prices) else 0.0
                    if price > 0:
                        # Buscar item existente
                        existing_item = PricelistItem.search([
                            ('pricelist_id', '=', pricelist.id),
                            ('product_tmpl_id', '=', product.id),
                        ], limit=1)

                        item_vals = {
                            'pricelist_id': pricelist.id,
                            'product_tmpl_id': product.id,
                            'applied_on': '1_product',
                            'compute_price': 'fixed',
                            'fixed_price': price,
                        }

                        if existing_item:
                            existing_item.write(item_vals)
                        else:
                            PricelistItem.create(item_vals)

                # Commit cada 100 productos para no perder progreso
                if (created + updated) % 100 == 0:
                    self.env.cr.commit()

            except Exception as e:
                errors.append(f'Fila {row_idx}: {str(e)}')
                if len(errors) > 50:
                    errors.append('... (demasiados errores, se detuvo el registro)')
                    break

        # Commit final
        self.env.cr.commit()

        # Generar mensaje de resultado
        result_parts = [
            f'Importación completada:',
            f'  - Productos creados: {created}',
            f'  - Productos actualizados: {updated}',
            f'  - Productos omitidos: {skipped}',
        ]
        if errors:
            result_parts.append(f'\nErrores ({len(errors)}):')
            for err in errors[:20]:
                result_parts.append(f'  - {err}')

        self.write({
            'result_message': '\n'.join(result_parts),
            'state': 'done',
        })

        _logger.info(
            'Importación Excel: %d creados, %d actualizados, %d omitidos, %d errores',
            created, updated, skipped, len(errors),
        )

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # ==================================================================
    # MANTENIMIENTO: archivar / restaurar / borrar productos importados
    # ==================================================================
    # Estas operaciones SOLO afectan productos con es_ferreteria=True y
    # SOLO excluyen aquellos con stock real (qty_available > 0) por
    # seguridad. Restringidas al grupo manager desde la vista.

    def _domain_productos_ferreteria_sin_stock(self, active):
        """Productos de ferreteria, sin stock real, en estado activo dado."""
        return [
            ('es_ferreteria', '=', True),
            ('active', '=', active),
            ('qty_available', '<=', 0),
        ]

    def action_archivar_productos(self):
        """Archiva (active=False) todos los productos de ferreteria sin stock."""
        self.ensure_one()
        Product = self.env['product.template']
        productos = Product.search(
            self._domain_productos_ferreteria_sin_stock(active=True)
        )
        con_stock = Product.search([
            ('es_ferreteria', '=', True),
            ('active', '=', True),
            ('qty_available', '>', 0),
        ])
        total = len(productos)
        productos.write({'active': False})

        msg = [
            'Archivado completado:',
            f'  - Productos archivados: {total}',
            f'  - Omitidos por tener stock > 0: {len(con_stock)}',
            '',
            'Los productos archivados pueden restaurarse con el boton',
            '"Restaurar archivados".',
        ]
        self.write({'result_message': '\n'.join(msg), 'state': 'done'})
        _logger.info(
            'Archivado de productos ferreteria: %d archivados, %d omitidos por stock',
            total, len(con_stock),
        )
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_restaurar_productos(self):
        """Restaura (active=True) productos de ferreteria archivados."""
        self.ensure_one()
        Product = self.env['product.template'].with_context(active_test=False)
        productos = Product.search([
            ('es_ferreteria', '=', True),
            ('active', '=', False),
        ])
        total = len(productos)
        productos.write({'active': True})

        msg = [
            'Restauracion completada:',
            f'  - Productos restaurados: {total}',
        ]
        self.write({'result_message': '\n'.join(msg), 'state': 'done'})
        _logger.info('Restaurados %d productos ferreteria', total)
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_borrar_productos(self):
        """Borra definitivamente productos sin stock. Si fallan por
        integridad referencial (movimientos, ventas, etc.), los archiva
        como fallback para no dejarlos a medias."""
        self.ensure_one()
        Product = self.env['product.template'].with_context(active_test=False)
        productos = Product.search(
            self._domain_productos_ferreteria_sin_stock(active=True)
        )
        # Tambien intentamos borrar los archivados
        archivados = Product.search([
            ('es_ferreteria', '=', True),
            ('active', '=', False),
            ('qty_available', '<=', 0),
        ])
        candidatos = productos | archivados
        con_stock = Product.search([
            ('es_ferreteria', '=', True),
            ('qty_available', '>', 0),
        ])

        borrados = 0
        archivados_fallback = 0
        # Intentamos borrar uno por uno, capturando errores de FK
        for producto in candidatos:
            try:
                with self.env.cr.savepoint():
                    producto.unlink()
                    borrados += 1
            except Exception as e:
                # No se pudo borrar (movimientos, etc.) -> archivar
                _logger.info(
                    'No se pudo borrar producto %s (id=%d): %s. Archivando.',
                    producto.default_code or producto.name, producto.id, e,
                )
                try:
                    producto.write({'active': False})
                    archivados_fallback += 1
                except Exception:
                    pass

        msg = [
            'Borrado definitivo completado:',
            f'  - Productos borrados: {borrados}',
            f'  - Archivados como fallback (tenian referencias): {archivados_fallback}',
            f'  - Omitidos por tener stock > 0: {len(con_stock)}',
        ]
        self.write({'result_message': '\n'.join(msg), 'state': 'done'})
        _logger.warning(
            'Borrado definitivo: %d borrados, %d archivados fallback, %d omitidos por stock',
            borrados, archivados_fallback, len(con_stock),
        )
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
